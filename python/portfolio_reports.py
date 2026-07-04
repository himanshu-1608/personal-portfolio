"""Portfolio analytics builders: per-trade P&L, MTF funding P&L, and the
day-by-day portfolio timeline. All pure functions of the input CSVs plus the
in-memory price rows; no network access. Split out of fetch_yahoo_prices.py so
the calculation logic is easy to locate and reason about in isolation.
"""

from __future__ import annotations

import csv
import re
import sys
from datetime import datetime
from typing import Dict, List, Sequence, Tuple

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover
    load_workbook = None  # type: ignore

from portfolio_core import (
    ALL_LEDGER_FILE,
    CONTRACT_NOTES_FILE,
    DATE_COLUMN,
    DAY_COUNT,
    MTF_PNL_HEADERS,
    MTF_PNL_OUTPUT_FILE,
    PNL_HEADERS,
    PNL_OUTPUT_FILE,
    PORTFOLIO_TIMELINE_FILE,
    PORTFOLIO_TIMELINE_HEADERS,
    STOCK_COLUMN,
    TRADEBOOK_FILE,
    format_number,
    format_quantity,
    load_latest_market_data,
    normalize_symbol,
    parse_amount,
)


def parse_contract_notes_charge_data() -> Dict[str, Dict[str, float]]:
    if load_workbook is None or not CONTRACT_NOTES_FILE.exists():
        return {}

    charge_labels = {
        "Taxable value of Supply (Brokerage) (₹)",
        "Exchange transaction charges (₹)",
        "Clearing charges (₹)",
        "CGST (@9% of Brok, SEBI, Trans & Clearing Charges) (₹)",
        "SGST (@9% of Brok, SEBI, Trans & Clearing Charges) (₹)",
        "IGST (@18% of Brok, SEBI, Trans & Clearing Charges) (₹)",
        "Securities transaction tax (₹)",
        "SEBI turnover fees (₹)",
        "Stamp duty (₹)",
        "NSE Investor Protection Fund Trust (₹)",
    }
    charge_data_by_date: Dict[str, Dict[str, float]] = {}

    workbook = load_workbook(CONTRACT_NOTES_FILE, read_only=True, data_only=False)
    for sheet_name in workbook.sheetnames:
        try:
            sheet_date = datetime.strptime(sheet_name.strip(), "%d-%m-%Y").date().isoformat()
        except ValueError:
            continue
        worksheet = workbook[sheet_name]
        symbol_sell_value: Dict[str, float] = {}
        total_charges = 0.0

        in_trade_rows = False
        for row_number in range(1, 600):
            first_cell = worksheet.cell(row=row_number, column=1).value
            first_text = str(first_cell or "").strip()
            if first_text == "Order No.":
                in_trade_rows = True
                continue

            if in_trade_rows:
                if first_text.startswith("Pay in/Pay out obligation"):
                    in_trade_rows = False
                else:
                    side = str(worksheet.cell(row=row_number, column=6).value or "").strip().upper()
                    contract_desc = str(worksheet.cell(row=row_number, column=5).value or "").strip()
                    net_total = worksheet.cell(row=row_number, column=13).value
                    if side == "S" and contract_desc:
                        symbol = normalize_symbol(contract_desc.split("-")[0])
                        if symbol:
                            try:
                                net_value = abs(float(net_total))
                            except (TypeError, ValueError):
                                net_value = 0.0
                            symbol_sell_value[symbol] = symbol_sell_value.get(symbol, 0.0) + net_value

            if first_text in charge_labels:
                # Contract note keeps charge values in "NCL-Cash" / "NET TOTAL" columns (H/K).
                raw_amount = worksheet.cell(row=row_number, column=8).value
                if raw_amount in (None, ""):
                    raw_amount = worksheet.cell(row=row_number, column=11).value
                try:
                    charge_amount = abs(float(raw_amount))
                except (TypeError, ValueError):
                    charge_amount = 0.0
                total_charges += charge_amount

        charge_data_by_date[sheet_date] = {
            "__total_charges__": total_charges,
            **symbol_sell_value,
        }

    return charge_data_by_date


def build_pnl_summary(report_rows: Sequence[Dict[str, str]]) -> List[Dict[str, str]]:
    if not TRADEBOOK_FILE.exists():
        return []

    market_data = load_latest_market_data(report_rows)
    recommendation_dates: Dict[str, str] = {}
    for row in report_rows:
        stock_code = (row.get(STOCK_COLUMN) or "").strip()
        recommendation_date = (row.get(DATE_COLUMN) or "").strip()
        base_symbol = normalize_symbol(stock_code.split(".")[0])
        if base_symbol and recommendation_date:
            recommendation_dates[base_symbol] = recommendation_date
    lots_by_symbol: Dict[str, List[Dict[str, float | str]]] = {}
    trades: List[Tuple[str, str, float, float, str, int]] = []

    with TRADEBOOK_FILE.open("r", newline="", encoding="utf-8-sig") as file:
        reader = csv.DictReader(file)
        required_columns = {"symbol", "trade_type", "quantity", "price", "trade_date"}
        if reader.fieldnames is None:
            raise ValueError(f"Tradebook file is empty: {TRADEBOOK_FILE}")
        missing_columns = required_columns.difference(reader.fieldnames)
        if missing_columns:
            raise ValueError(
                "Tradebook file is missing required columns: "
                + ", ".join(sorted(missing_columns))
            )

        for row_number, row in enumerate(reader, start=2):
            symbol = normalize_symbol((row.get("symbol") or "").strip())
            trade_type = (row.get("trade_type") or "").strip().lower()
            trade_date = (row.get("trade_date") or "").strip()
            if not symbol or trade_type not in {"buy", "sell"}:
                raise ValueError(
                    f"Tradebook row {row_number} has invalid symbol or trade_type."
                )
            if not trade_date:
                raise ValueError(f"Tradebook row {row_number} is missing trade_date.")

            try:
                quantity = float((row.get("quantity") or "").strip())
                price = float((row.get("price") or "").strip())
            except ValueError as exc:
                raise ValueError(
                    f"Tradebook row {row_number} has invalid numeric quantity or price."
                ) from exc

            sort_timestamp = (row.get("order_execution_time") or "").strip() or trade_date
            trades.append((sort_timestamp, symbol, quantity, price, trade_type, row_number))

    trades.sort(key=lambda item: (item[0], item[5]))

    for _sort_time, symbol, quantity, price, trade_type, row_number in trades:
        symbol_lots = lots_by_symbol.setdefault(symbol, [])

        if trade_type == "buy":
            trade_date = _sort_time[:10]
            lot = next(
                (
                    entry
                    for entry in symbol_lots
                    if str(entry["trade_date"]) == trade_date
                ),
                None,
            )
            if lot is None:
                lot = {
                    "trade_date": trade_date,
                    "buy_quantity": 0.0,
                    "buy_value": 0.0,
                    "sell_quantity": 0.0,
                    "sell_value": 0.0,
                    "position_quantity": 0.0,
                    "position_cost": 0.0,
                    "realized_pnl": 0.0,
                    "charges_taxes_others": 0.0,
                    "realized_cost_basis": 0.0,
                }
                symbol_lots.append(lot)

            lot["buy_quantity"] = float(lot["buy_quantity"]) + quantity
            lot["buy_value"] = float(lot["buy_value"]) + (quantity * price)
            lot["position_quantity"] = float(lot["position_quantity"]) + quantity
            lot["position_cost"] = float(lot["position_cost"]) + (quantity * price)
            continue

        remaining_to_sell = quantity
        for lot in symbol_lots:
            lot_position_quantity = float(lot["position_quantity"])
            if lot_position_quantity <= 0:
                continue

            matched_quantity = min(remaining_to_sell, lot_position_quantity)
            average_cost = float(lot["position_cost"]) / lot_position_quantity if lot_position_quantity else 0.0

            lot["sell_quantity"] = float(lot["sell_quantity"]) + matched_quantity
            lot["sell_value"] = float(lot["sell_value"]) + (matched_quantity * price)
            lot["realized_pnl"] = float(lot["realized_pnl"]) + (matched_quantity * (price - average_cost))
            lot["realized_cost_basis"] = float(lot["realized_cost_basis"]) + (matched_quantity * average_cost)
            lot["position_quantity"] = lot_position_quantity - matched_quantity
            lot["position_cost"] = float(lot["position_cost"]) - (matched_quantity * average_cost)

            remaining_to_sell -= matched_quantity
            if remaining_to_sell <= 0:
                break

        # If sells exceed open lots, keep behavior non-short: ignore unmatched remainder.
        if remaining_to_sell > 0:
            print(
                f"Warning: Ignored unmatched sell quantity ({remaining_to_sell:.4f}) for {symbol} at row {row_number}.",
                file=sys.stderr,
            )

    contract_note_charges = parse_contract_notes_charge_data()
    if contract_note_charges:
        unallocated_contract_note_charges = 0.0
        for charge_date, payload in contract_note_charges.items():
            total_charges = float(payload.get("__total_charges__", 0.0))
            if total_charges <= 0:
                continue

            symbol_weights = {
                symbol: float(value)
                for symbol, value in payload.items()
                if symbol != "__total_charges__" and float(value) > 0
            }
            total_symbol_weight = sum(symbol_weights.values())

            date_lots_all_symbols: List[Tuple[str, Dict[str, float | str]]] = []
            for symbol, symbol_lots in lots_by_symbol.items():
                for lot in symbol_lots:
                    if str(lot["trade_date"]) != charge_date:
                        continue
                    if float(lot["sell_value"]) <= 0:
                        continue
                    date_lots_all_symbols.append((symbol, lot))

            if not date_lots_all_symbols:
                unallocated_contract_note_charges += total_charges
                continue

            per_symbol_charge: Dict[str, float] = {}
            if total_symbol_weight > 0:
                running_allocated = 0.0
                symbols = sorted(symbol_weights.keys())
                for index, symbol in enumerate(symbols):
                    symbol_weight = symbol_weights[symbol]
                    if index == len(symbols) - 1:
                        allocated_charge = total_charges - running_allocated
                    else:
                        allocated_charge = total_charges * (symbol_weight / total_symbol_weight)
                        running_allocated += allocated_charge
                    per_symbol_charge[symbol] = per_symbol_charge.get(symbol, 0.0) + allocated_charge
            else:
                realized_value_on_date = sum(float(lot["sell_value"]) for _, lot in date_lots_all_symbols)
                if realized_value_on_date <= 0:
                    continue
                running_allocated = 0.0
                for index, (symbol, lot) in enumerate(date_lots_all_symbols):
                    sell_value = float(lot["sell_value"])
                    if index == len(date_lots_all_symbols) - 1:
                        allocated_charge = total_charges - running_allocated
                    else:
                        allocated_charge = total_charges * (sell_value / realized_value_on_date)
                        running_allocated += allocated_charge
                    lot["charges_taxes_others"] = float(lot["charges_taxes_others"]) + allocated_charge
                continue

            for symbol, symbol_charge in per_symbol_charge.items():
                symbol_lots = [
                    lot
                    for lot_symbol, lot in date_lots_all_symbols
                    if lot_symbol == symbol
                ]
                if not symbol_lots:
                    unallocated_contract_note_charges += symbol_charge
                    continue
                symbol_sell_value = sum(float(lot["sell_value"]) for lot in symbol_lots)
                if symbol_sell_value <= 0:
                    unallocated_contract_note_charges += symbol_charge
                    continue
                running_allocated = 0.0
                for index, lot in enumerate(symbol_lots):
                    sell_value = float(lot["sell_value"])
                    if index == len(symbol_lots) - 1:
                        allocated_charge = symbol_charge - running_allocated
                    else:
                        allocated_charge = symbol_charge * (sell_value / symbol_sell_value)
                        running_allocated += allocated_charge
                    lot["charges_taxes_others"] = float(lot["charges_taxes_others"]) + allocated_charge

        if unallocated_contract_note_charges > 0:
            realized_lots_all_symbols = [
                lot
                for symbol_lots in lots_by_symbol.values()
                for lot in symbol_lots
                if float(lot["sell_value"]) > 0
            ]
            realized_total_sell_value = sum(float(lot["sell_value"]) for lot in realized_lots_all_symbols)
            if realized_total_sell_value > 0:
                allocated = 0.0
                for index, lot in enumerate(realized_lots_all_symbols):
                    sell_value = float(lot["sell_value"])
                    if index == len(realized_lots_all_symbols) - 1:
                        lot_charge = unallocated_contract_note_charges - allocated
                    else:
                        lot_charge = unallocated_contract_note_charges * (sell_value / realized_total_sell_value)
                        allocated += lot_charge
                    lot["charges_taxes_others"] = float(lot["charges_taxes_others"]) + lot_charge
    else:
        sale_charge_events: List[Tuple[str, float]] = []
        generic_charge_pool = 0.0
        if ALL_LEDGER_FILE.exists():
            with ALL_LEDGER_FILE.open("r", newline="", encoding="utf-8-sig") as file:
                reader = csv.DictReader(file)
                for row in reader:
                    particulars = (row.get("particulars") or "").strip()
                    particulars_lower = particulars.lower()
                    debit_amount = parse_amount(row.get("debit") or "")
                    if debit_amount <= 0:
                        continue

                    # Inclusion approach: only count debits that are explicitly known
                    # brokerage/exchange charges. Unknown entries are ignored so new
                    # ledger categories never silently inflate the charge pool.
                    CHARGE_INCLUDE_TOKENS = (
                        "dp charges",
                        "brokerage",
                        "stamp duty",
                        "stt",
                        "securities transaction tax",
                        "sebi",
                        "exchange transaction charge",
                        "clearing charge",
                        "cgst",
                        "sgst",
                        "igst",
                    )
                    is_known_charge = any(token in particulars_lower for token in CHARGE_INCLUDE_TOKENS)
                    if not is_known_charge:
                        continue

                    matched_sale_charge = re.search(r"sale of ([A-Za-z0-9._-]+)", particulars, flags=re.IGNORECASE)
                    if matched_sale_charge:
                        charge_symbol = normalize_symbol(matched_sale_charge.group(1))
                        if charge_symbol:
                            sale_charge_events.append((charge_symbol, debit_amount))
                            continue

                    generic_charge_pool += debit_amount

        for symbol, charge_amount in sale_charge_events:
            symbol_lots = lots_by_symbol.get(symbol, [])
            realized_lots = [lot for lot in symbol_lots if float(lot["sell_value"]) > 0]
            total_sell_value = sum(float(lot["sell_value"]) for lot in realized_lots)
            if total_sell_value <= 0:
                generic_charge_pool += charge_amount
                continue
            allocated = 0.0
            for index, lot in enumerate(realized_lots):
                sell_value = float(lot["sell_value"])
                if index == len(realized_lots) - 1:
                    lot_charge = charge_amount - allocated
                else:
                    lot_charge = charge_amount * (sell_value / total_sell_value)
                    allocated += lot_charge
                lot["charges_taxes_others"] = float(lot["charges_taxes_others"]) + lot_charge

        realized_lots_all_symbols = [
            lot
            for symbol_lots in lots_by_symbol.values()
            for lot in symbol_lots
            if float(lot["sell_value"]) > 0
        ]
        realized_total_sell_value = sum(float(lot["sell_value"]) for lot in realized_lots_all_symbols)
        if generic_charge_pool > 0 and realized_total_sell_value > 0:
            allocated = 0.0
            for index, lot in enumerate(realized_lots_all_symbols):
                sell_value = float(lot["sell_value"])
                if index == len(realized_lots_all_symbols) - 1:
                    lot_charge = generic_charge_pool - allocated
                else:
                    lot_charge = generic_charge_pool * (sell_value / realized_total_sell_value)
                    allocated += lot_charge
                lot["charges_taxes_others"] = float(lot["charges_taxes_others"]) + lot_charge

    sortable_rows: List[Tuple[str, str, Dict[str, str]]] = []
    for symbol, symbol_lots in lots_by_symbol.items():
        for lot in symbol_lots:
            position_quantity = float(lot["position_quantity"])
            position_cost = float(lot["position_cost"])
            buy_quantity = float(lot["buy_quantity"])
            buy_value = float(lot["buy_value"])
            sell_quantity = float(lot["sell_quantity"])
            sell_value = float(lot["sell_value"])
            realized_pnl = float(lot["realized_pnl"])
            charges_taxes_others = float(lot["charges_taxes_others"])
            net_realized_pnl = realized_pnl - charges_taxes_others

            average_buy_price = (buy_value / buy_quantity) if buy_quantity else 0.0
            average_sell_price = (sell_value / sell_quantity) if sell_quantity else 0.0
            open_average_cost = (position_cost / position_quantity) if position_quantity else 0.0

            market_entry = market_data.get(symbol, {})
            latest_market_price = float(market_entry.get("latest_market_price", "0") or 0.0)
            latest_market_date = str(market_entry.get("latest_market_date", "") or "")
            matched_report_symbol = str(market_entry.get("report_symbol", symbol) or symbol)
            unrealized_pnl = (
                position_quantity * (latest_market_price - open_average_cost)
                if position_quantity and latest_market_price
                else 0.0
            )
            total_pnl = net_realized_pnl + unrealized_pnl
            total_return_pct = ((total_pnl / buy_value) * 100) if buy_value else None

            row = {
                "Stock Code/Name": symbol,
                "Trade Date": str(lot["trade_date"]),
                "Matched Report Symbol": matched_report_symbol,
                "Buy Quantity": format_quantity(buy_quantity),
                "Buy Value": format_number(buy_value),
                "Sell Quantity": format_quantity(sell_quantity),
                "Sell Value": format_number(sell_value),
                "Net Quantity": format_quantity(position_quantity),
                "Average Buy Price": format_number(average_buy_price),
                "Average Sell Price": format_number(average_sell_price),
                "Open Average Cost": format_number(open_average_cost),
                "Latest Market Price": format_number(latest_market_price) if latest_market_price else "",
                "Latest Market Date": latest_market_date,
                "Realized P&L": format_number(realized_pnl),
                "Charges, Taxes, Others": format_number(charges_taxes_others),
                "Net Realized P&L": format_number(net_realized_pnl),
                "Unrealized P&L": format_number(unrealized_pnl),
                "Total P&L": format_number(total_pnl),
                "Return %": f"{total_return_pct:.2f}%" if total_return_pct is not None else "",
            }
            sort_date = recommendation_dates.get(symbol, "9999-12-31")
            trade_date = str(lot["trade_date"])
            sortable_rows.append((sort_date, symbol, trade_date, row))

    sortable_rows.sort(key=lambda item: (item[0], item[1], item[2]))
    return [row for _date, _symbol, _trade_date, row in sortable_rows]


def build_portfolio_timeline(report_rows: Sequence[Dict[str, str]]) -> List[Dict[str, str]]:
    """Compute real day-by-day portfolio components, split by realized/unrealized
    and by full-value/MTF-funding, so the dashboard can render any filter combo.

    Per date we emit, for both realized (sold-by-date) and unrealized (still-held)
    buckets: full buy value, current value, and the zerodha (MTF) funding portion.
    The frontend derives the 6 filter combinations:
      - basis "all holdings"  -> use full value (r_invested / u_invested, r_value / u_value)
      - basis "personal"      -> subtract zerodha funding (full - zerodha)
      - filter realized/unrealized/all -> pick r_*, u_*, or their sum
    """
    import bisect

    if not TRADEBOOK_FILE.exists():
        return []

    # Per-buy-date MTF margin ratio. ratio = your_funding / buy_value.
    # Cash (non-MTF) buy dates are absent -> treated as ratio 1.0 (zero zerodha funding).
    margin_ratio_by_date = build_mtf_margin_ratio_by_date()

    # Build price lookup: {base_symbol: {iso_date: price}}
    price_by_symbol: Dict[str, Dict[str, float]] = {}
    for row in report_rows:
        stock_code = (row.get(STOCK_COLUMN) or "").strip()
        base_symbol = normalize_symbol(stock_code.split(".")[0])
        if not base_symbol:
            continue
        sym_prices = price_by_symbol.setdefault(base_symbol, {})
        for index in range(1, DAY_COUNT + 1):
            row_date = (row.get(f"Date {index}") or "").strip()
            row_price = (row.get(f"Day {index} Price") or "").strip()
            if row_date and row_price:
                try:
                    sym_prices[row_date] = float(row_price)
                except ValueError:
                    pass

    if not price_by_symbol:
        return []

    # Read tradebook into chronological trade events (mirror FIFO of build_pnl_summary).
    trades: List[Tuple[str, str, float, float, str]] = []  # (sort_time, symbol, qty, price, trade_type)
    with TRADEBOOK_FILE.open("r", newline="", encoding="utf-8-sig") as file:
        reader = csv.DictReader(file)
        if reader.fieldnames is None:
            return []
        for row in reader:
            symbol = normalize_symbol((row.get("symbol") or "").strip())
            trade_type = (row.get("trade_type") or "").strip().lower()
            quantity = parse_amount(row.get("quantity") or "")
            price = parse_amount(row.get("price") or "")
            trade_date = (row.get("trade_date") or "").strip()
            if not symbol or trade_type not in {"buy", "sell"} or not trade_date:
                continue
            if quantity <= 0 or price <= 0:
                continue
            sort_time = (row.get("order_execution_time") or "").strip() or trade_date
            trades.append((sort_time, symbol, quantity, price, trade_type))

    if not trades:
        return []

    trades.sort(key=lambda t: t[0])

    # Build lots aggregated by (symbol, buy_date). FIFO sells record per-lot sell events
    # with their dates so we can reconstruct realized/unrealized split at any past date.
    lots_by_symbol: Dict[str, List[Dict[str, object]]] = {}
    for sort_time, symbol, quantity, price, trade_type in trades:
        symbol_lots = lots_by_symbol.setdefault(symbol, [])
        if trade_type == "buy":
            trade_date = sort_time[:10]
            lot = next((l for l in symbol_lots if l["trade_date"] == trade_date), None)
            if lot is None:
                lot = {
                    "trade_date": trade_date,
                    "buy_quantity": 0.0,
                    "buy_value": 0.0,
                    "zerodha_value": 0.0,
                    "position_quantity": 0.0,
                    "sells": [],  # list of (sell_date, qty, sell_price)
                }
                symbol_lots.append(lot)
            ratio = margin_ratio_by_date.get(trade_date, 1.0)
            buy_value = quantity * price
            lot["buy_quantity"] = float(lot["buy_quantity"]) + quantity
            lot["buy_value"] = float(lot["buy_value"]) + buy_value
            lot["zerodha_value"] = float(lot["zerodha_value"]) + buy_value * (1.0 - ratio)
            lot["position_quantity"] = float(lot["position_quantity"]) + quantity
            continue

        # sell: FIFO across this symbol's lots in chronological order
        sell_date = sort_time[:10]
        remaining = quantity
        for lot in symbol_lots:
            open_qty = float(lot["position_quantity"])
            if open_qty <= 0:
                continue
            matched = min(remaining, open_qty)
            lot["position_quantity"] = open_qty - matched
            lot["sells"].append((sell_date, matched, price))
            remaining -= matched
            if remaining <= 0:
                break

    # All market dates across all stocks in the price data
    all_market_dates = sorted({d for sym_prices in price_by_symbol.values() for d in sym_prices})
    if not all_market_dates:
        return []

    sorted_dates_by_symbol: Dict[str, List[str]] = {
        sym: sorted(dates.keys()) for sym, dates in price_by_symbol.items()
    }

    def price_on(symbol: str, market_date: str, fallback: float) -> float:
        sym_dates = sorted_dates_by_symbol.get(symbol)
        sym_prices = price_by_symbol.get(symbol)
        if not sym_dates or not sym_prices:
            return fallback
        pos = bisect.bisect_right(sym_dates, market_date)
        if pos == 0:
            return fallback
        return sym_prices[sym_dates[pos - 1]]

    timeline: List[Dict[str, str]] = []
    for market_date in all_market_dates:
        r_invested = r_value = r_zerodha = 0.0
        u_invested = u_value = u_zerodha = 0.0

        for symbol, symbol_lots in lots_by_symbol.items():
            for lot in symbol_lots:
                if lot["trade_date"] > market_date:
                    continue  # lot not bought yet
                buy_qty = float(lot["buy_quantity"])
                if buy_qty <= 0:
                    continue
                buy_price_ps = float(lot["buy_value"]) / buy_qty
                zerodha_ps = float(lot["zerodha_value"]) / buy_qty

                sold_qty = 0.0
                realized_value = 0.0
                for sell_date, qty, sell_price in lot["sells"]:
                    if sell_date <= market_date:
                        sold_qty += qty
                        realized_value += qty * sell_price
                open_qty = buy_qty - sold_qty

                if sold_qty > 0:
                    r_invested += sold_qty * buy_price_ps
                    r_value += realized_value
                    r_zerodha += sold_qty * zerodha_ps
                if open_qty > 0:
                    u_invested += open_qty * buy_price_ps
                    u_value += open_qty * price_on(symbol, market_date, buy_price_ps)
                    u_zerodha += open_qty * zerodha_ps

        if r_invested <= 0 and u_invested <= 0:
            continue

        timeline.append({
            "date": market_date,
            "r_invested": f"{r_invested:.2f}",
            "r_value": f"{r_value:.2f}",
            "r_zerodha": f"{r_zerodha:.2f}",
            "u_invested": f"{u_invested:.2f}",
            "u_value": f"{u_value:.2f}",
            "u_zerodha": f"{u_zerodha:.2f}",
        })

    return timeline


def write_portfolio_timeline_output(rows: Sequence[Dict[str, str]]) -> None:
    with PORTFOLIO_TIMELINE_FILE.open("w", newline="", encoding="utf-8") as file:
        writer = csv.DictWriter(file, fieldnames=PORTFOLIO_TIMELINE_HEADERS)
        writer.writeheader()
        writer.writerows(rows)


def write_pnl_output(rows: Sequence[Dict[str, str]]) -> None:
    with PNL_OUTPUT_FILE.open("w", newline="", encoding="utf-8") as file:
        writer = csv.DictWriter(file, fieldnames=PNL_HEADERS)
        writer.writeheader()
        writer.writerows(rows)


def _mtf_symbols_from_ledger() -> set:
    """Symbols that were pledged/unpledged = the ones actually bought on MTF.

    Zerodha names the symbol directly in "MTF pledge charges for <SYMBOL>" /
    "MTF unpledge charges for <SYMBOL>" rows, so this is a reliable per-symbol
    MTF signal independent of the (now-missing) gross-obligation rows."""
    symbols: set = set()
    if not ALL_LEDGER_FILE.exists():
        return symbols
    with ALL_LEDGER_FILE.open("r", newline="", encoding="utf-8-sig") as file:
        reader = csv.DictReader(file)
        for row in reader:
            particulars = (row.get("particulars") or "").strip()
            lower = particulars.lower()
            if "mtf pledge charges for " not in lower and "mtf unpledge charges for " not in lower:
                continue
            matched = re.search(r"charges for\s+([A-Za-z0-9._-]+)", particulars, flags=re.IGNORECASE)
            if matched:
                symbols.add(normalize_symbol(matched.group(1)))
    return symbols


def _mtf_buy_value_by_date(mtf_symbols: set) -> Dict[str, float]:
    """Total MTF buy value (qty x price) per trade date, from the tradebook,
    restricted to the pledged (MTF) symbols. This reconstructs the gross MTF
    obligation now that Zerodha no longer emits "Gross obligation for MTF" rows."""
    buy_value_by_date: Dict[str, float] = {}
    if not TRADEBOOK_FILE.exists() or not mtf_symbols:
        return buy_value_by_date
    with TRADEBOOK_FILE.open("r", newline="", encoding="utf-8-sig") as file:
        reader = csv.DictReader(file)
        for row in reader:
            if (row.get("trade_type") or "").strip().lower() != "buy":
                continue
            symbol = normalize_symbol((row.get("symbol") or "").strip())
            trade_date = (row.get("trade_date") or "").strip()
            if not symbol or not trade_date or symbol not in mtf_symbols:
                continue
            quantity = parse_amount(row.get("quantity") or "")
            price = parse_amount(row.get("price") or "")
            if quantity <= 0 or price <= 0:
                continue
            buy_value_by_date[trade_date] = buy_value_by_date.get(trade_date, 0.0) + quantity * price
    return buy_value_by_date


def build_mtf_margin_ratio_by_date() -> Dict[str, float]:
    if not ALL_LEDGER_FILE.exists():
        return {}

    # "Initial margin charged for MTF" posts on the trade date and marks which
    # dates had MTF buys (value = your-money portion of the buy).
    initial_margin_by_date: Dict[str, float] = {}
    with ALL_LEDGER_FILE.open("r", newline="", encoding="utf-8-sig") as file:
        reader = csv.DictReader(file)
        for row in reader:
            particulars = (row.get("particulars") or "").strip().lower()
            posting_date = (row.get("posting_date") or "").strip()
            if not posting_date:
                continue
            if particulars != "initial margin charged for mtf":
                continue
            initial_margin_by_date[posting_date] = initial_margin_by_date.get(posting_date, 0.0) + parse_amount(
                row.get("debit") or ""
            )

    # Gross MTF obligation (total buy value) per date, from the tradebook.
    # Zerodha used to emit an explicit "Gross obligation for MTF" ledger row but
    # has stopped, and the rows it did emit were sometimes partial (e.g.
    # SANGHVIMOV 2026-06-09 showed a 10,363 obligation against a 124,328 buy,
    # producing an impossible margin ratio > 1 and negative broker funding). The
    # tradebook buy value (qty x price for the pledged MTF symbols) is the true
    # obligation.
    buy_value_by_date = _mtf_buy_value_by_date(_mtf_symbols_from_ledger())

    ratios: Dict[str, float] = {}
    for posting_date, margin_value in initial_margin_by_date.items():
        gross_value = buy_value_by_date.get(posting_date, 0.0)
        if gross_value <= 0:
            continue
        # Margin can't exceed the buy value; clamp so a lagging/partial margin
        # row never produces a >1 ratio (which would flip funding negative).
        ratios[posting_date] = min(margin_value / gross_value, 1.0)
    return ratios


def build_mtf_pnl_summary(report_rows: Sequence[Dict[str, str]]) -> List[Dict[str, str]]:
    if not TRADEBOOK_FILE.exists():
        return []

    margin_ratio_by_date = build_mtf_margin_ratio_by_date()
    if not margin_ratio_by_date:
        return []

    # A buy is MTF only if its symbol was actually pledged on MTF. Sharing a
    # trade date with an MTF buy is not enough (cash buys can land the same day),
    # so gate the is_mtf flag on this per-symbol set rather than the date alone.
    mtf_symbols = _mtf_symbols_from_ledger()

    market_data = load_latest_market_data(report_rows)
    lots_by_symbol: Dict[str, List[Dict[str, float | str]]] = {}
    events: List[Tuple[str, str, Dict[str, float | str], int]] = []

    with TRADEBOOK_FILE.open("r", newline="", encoding="utf-8-sig") as file:
        reader = csv.DictReader(file)
        required_columns = {"symbol", "trade_type", "quantity", "price", "trade_date"}
        if reader.fieldnames is None:
            raise ValueError(f"Tradebook file is empty: {TRADEBOOK_FILE}")
        missing_columns = required_columns.difference(reader.fieldnames)
        if missing_columns:
            raise ValueError(
                "Tradebook file is missing required columns: "
                + ", ".join(sorted(missing_columns))
            )

        for row_number, row in enumerate(reader, start=2):
            symbol = normalize_symbol((row.get("symbol") or "").strip())
            trade_type = (row.get("trade_type") or "").strip().lower()
            trade_date = (row.get("trade_date") or "").strip()
            if not symbol or trade_type not in {"buy", "sell"} or not trade_date:
                continue

            quantity = parse_amount(row.get("quantity") or "")
            price = parse_amount(row.get("price") or "")
            if quantity <= 0 or price <= 0:
                continue

            sort_timestamp = (row.get("order_execution_time") or "").strip() or trade_date
            events.append(
                (
                    sort_timestamp,
                    "trade",
                    {
                        "symbol": symbol,
                        "quantity": quantity,
                        "price": price,
                        "trade_type": trade_type,
                        "trade_date": trade_date,
                    },
                    row_number,
                )
            )

    if ALL_LEDGER_FILE.exists():
        with ALL_LEDGER_FILE.open("r", newline="", encoding="utf-8-sig") as file:
            reader = csv.DictReader(file)
            for row_number, row in enumerate(reader, start=2):
                particulars = (row.get("particulars") or "").strip()
                particulars_lower = particulars.lower()
                posting_date = (row.get("posting_date") or "").strip()
                if not posting_date:
                    continue

                if "interest for mtf funded value on " in particulars_lower:
                    matched = re.search(r"on\s+(\d{4}-\d{2}-\d{2})", particulars)
                    accrual_date = matched.group(1) if matched else posting_date
                    debit_value = parse_amount(row.get("debit") or "")
                    if debit_value <= 0:
                        continue
                    events.append(
                        (
                            f"{accrual_date}T23:59:59",
                            "interest",
                            {"amount": debit_value},
                            row_number,
                        )
                    )
                    continue

                if "mtf pledge charges for " in particulars_lower or "mtf unpledge charges for " in particulars_lower:
                    debit_value = parse_amount(row.get("debit") or "")
                    if debit_value <= 0:
                        continue
                    matched_symbol = re.search(r"for\s+([A-Za-z0-9._-]+)", particulars, flags=re.IGNORECASE)
                    if not matched_symbol:
                        continue
                    events.append(
                        (
                            f"{posting_date}T23:59:59",
                            "pledge",
                            {
                                "symbol": normalize_symbol(matched_symbol.group(1)),
                                "amount": debit_value,
                            },
                            row_number,
                        )
                    )

    events.sort(key=lambda item: (item[0], item[3]))

    def allocate_amount_across_lots(target_lots: Sequence[Dict[str, float | str]], amount: float, field: str) -> None:
        total_open_funding = sum(float(lot["open_zerodha_funding"]) for lot in target_lots if float(lot["open_zerodha_funding"]) > 0)
        if amount <= 0 or total_open_funding <= 0:
            return

        running_allocated = 0.0
        open_lots = [lot for lot in target_lots if float(lot["open_zerodha_funding"]) > 0]
        for index, lot in enumerate(open_lots):
            open_funding = float(lot["open_zerodha_funding"])
            if index == len(open_lots) - 1:
                allocation = amount - running_allocated
            else:
                allocation = amount * (open_funding / total_open_funding)
                running_allocated += allocation
            lot[field] = float(lot[field]) + allocation

    for _event_time, event_type, payload, _row_number in events:
        if event_type == "trade":
            symbol = str(payload["symbol"])
            quantity = float(payload["quantity"])
            price = float(payload["price"])
            trade_type = str(payload["trade_type"])
            trade_date = str(payload["trade_date"])
            symbol_lots = lots_by_symbol.setdefault(symbol, [])

            if trade_type == "buy":
                # Cash (non-MTF) buys still create a lot so FIFO sell-matching stays
                # aligned with the P&L summary; they carry zero funding and are excluded
                # from the final MTF rows via the is_mtf flag. Skipping them here would
                # mis-assign sell prices to MTF lots for mixed cash+MTF symbols.
                ratio = margin_ratio_by_date.get(trade_date)
                is_mtf = symbol in mtf_symbols and ratio is not None
                effective_ratio = ratio if ratio is not None else 1.0
                lot = next((entry for entry in symbol_lots if str(entry["trade_date"]) == trade_date), None)
                if lot is None:
                    lot = {
                        "trade_date": trade_date,
                        "is_mtf": is_mtf,
                        "buy_quantity": 0.0,
                        "buy_value": 0.0,
                        "your_funding": 0.0,
                        "zerodha_funding": 0.0,
                        "sell_quantity": 0.0,
                        "sell_value": 0.0,
                        "position_quantity": 0.0,
                        "position_cost": 0.0,
                        "open_zerodha_funding": 0.0,
                        "realized_pnl": 0.0,
                        "mtf_interest_cost": 0.0,
                        "mtf_pledge_charges": 0.0,
                    }
                    symbol_lots.append(lot)

                buy_value = quantity * price
                your_funding = buy_value * effective_ratio
                zerodha_funding = buy_value - your_funding
                lot["buy_quantity"] = float(lot["buy_quantity"]) + quantity
                lot["buy_value"] = float(lot["buy_value"]) + buy_value
                lot["your_funding"] = float(lot["your_funding"]) + your_funding
                lot["zerodha_funding"] = float(lot["zerodha_funding"]) + zerodha_funding
                lot["position_quantity"] = float(lot["position_quantity"]) + quantity
                lot["position_cost"] = float(lot["position_cost"]) + buy_value
                lot["open_zerodha_funding"] = float(lot["open_zerodha_funding"]) + zerodha_funding
                continue

            remaining_to_sell = quantity
            for lot in symbol_lots:
                lot_position_quantity = float(lot["position_quantity"])
                if lot_position_quantity <= 0:
                    continue

                matched_quantity = min(remaining_to_sell, lot_position_quantity)
                average_cost = float(lot["position_cost"]) / lot_position_quantity if lot_position_quantity else 0.0
                matched_value = matched_quantity * price

                lot["sell_quantity"] = float(lot["sell_quantity"]) + matched_quantity
                lot["sell_value"] = float(lot["sell_value"]) + matched_value
                lot["realized_pnl"] = float(lot["realized_pnl"]) + (matched_quantity * (price - average_cost))
                lot["position_quantity"] = lot_position_quantity - matched_quantity
                lot["position_cost"] = float(lot["position_cost"]) - (matched_quantity * average_cost)

                open_funding = float(lot["open_zerodha_funding"])
                funding_per_share = open_funding / lot_position_quantity if lot_position_quantity else 0.0
                lot["open_zerodha_funding"] = max(0.0, open_funding - (matched_quantity * funding_per_share))

                remaining_to_sell -= matched_quantity
                if remaining_to_sell <= 0:
                    break
            continue

        if event_type == "interest":
            open_lots = [
                lot
                for symbol_lots in lots_by_symbol.values()
                for lot in symbol_lots
                if float(lot["open_zerodha_funding"]) > 0
            ]
            allocate_amount_across_lots(open_lots, float(payload["amount"]), "mtf_interest_cost")
            continue

        if event_type == "pledge":
            symbol = str(payload["symbol"])
            symbol_lots = lots_by_symbol.get(symbol, [])
            open_symbol_lots = [lot for lot in symbol_lots if float(lot["open_zerodha_funding"]) > 0]
            allocate_amount_across_lots(open_symbol_lots, float(payload["amount"]), "mtf_pledge_charges")

    sortable_rows: List[Tuple[str, str, Dict[str, str]]] = []
    for symbol, symbol_lots in lots_by_symbol.items():
        for lot in symbol_lots:
            buy_quantity = float(lot["buy_quantity"])
            buy_value = float(lot["buy_value"])
            your_funding = float(lot["your_funding"])
            zerodha_funding = float(lot["zerodha_funding"])
            sell_quantity = float(lot["sell_quantity"])
            sell_value = float(lot["sell_value"])
            position_quantity = float(lot["position_quantity"])
            position_cost = float(lot["position_cost"])
            realized_pnl = float(lot["realized_pnl"])
            mtf_interest_cost = float(lot["mtf_interest_cost"])
            mtf_pledge_charges = float(lot["mtf_pledge_charges"])

            if buy_quantity <= 0:
                continue
            if not lot.get("is_mtf"):
                continue  # cash lot — tracked only for FIFO alignment, not an MTF row

            market_entry = market_data.get(symbol, {})
            latest_market_price = float(market_entry.get("latest_market_price", "0") or 0.0)
            latest_market_date = str(market_entry.get("latest_market_date", "") or "")
            matched_report_symbol = str(market_entry.get("report_symbol", symbol) or symbol)
            open_average_cost = (position_cost / position_quantity) if position_quantity else 0.0
            unrealized_pnl = (
                position_quantity * (latest_market_price - open_average_cost)
                if position_quantity and latest_market_price
                else 0.0
            )
            holding_pnl = realized_pnl + unrealized_pnl
            holding_return_pct = ((holding_pnl / buy_value) * 100) if buy_value else None
            funding_return_pct = ((holding_pnl / your_funding) * 100) if your_funding else None
            carrying_cost = mtf_interest_cost + mtf_pledge_charges
            net_funding_pnl = holding_pnl - carrying_cost
            net_funding_return_pct = ((net_funding_pnl / your_funding) * 100) if your_funding else None
            leverage = (buy_value / your_funding) if your_funding else None

            row = {
                "Stock Code/Name": symbol,
                "Trade Date": str(lot["trade_date"]),
                "Matched Report Symbol": matched_report_symbol,
                "Buy Quantity": format_quantity(buy_quantity),
                "Buy Value": format_number(buy_value),
                "Your Funding": format_number(your_funding),
                "Zerodha Funding": format_number(zerodha_funding),
                "Leverage X": f"{leverage:.2f}x" if leverage is not None else "",
                "Sell Quantity": format_quantity(sell_quantity),
                "Sell Value": format_number(sell_value),
                "Net Quantity": format_quantity(position_quantity),
                "Latest Market Price": format_number(latest_market_price) if latest_market_price else "",
                "Latest Market Date": latest_market_date,
                "Holding P&L": format_number(holding_pnl),
                "Holding Return %": f"{holding_return_pct:.2f}%" if holding_return_pct is not None else "",
                "Funding Return %": f"{funding_return_pct:.2f}%" if funding_return_pct is not None else "",
                "MTF Interest Cost": format_number(mtf_interest_cost),
                "MTF Pledge Charges": format_number(mtf_pledge_charges),
                "Total Carrying Cost": format_number(carrying_cost),
                "Net Funding P&L": format_number(net_funding_pnl),
                "Net Funding Return %": f"{net_funding_return_pct:.2f}%" if net_funding_return_pct is not None else "",
            }
            trade_date = str(lot["trade_date"])
            sortable_rows.append((trade_date, symbol, row))

    sortable_rows.sort(key=lambda item: (item[0], item[1]))
    return [row for _trade_date, _symbol, row in sortable_rows]


def write_mtf_pnl_output(rows: Sequence[Dict[str, str]]) -> None:
    with MTF_PNL_OUTPUT_FILE.open("w", newline="", encoding="utf-8") as file:
        writer = csv.DictWriter(file, fieldnames=MTF_PNL_HEADERS)
        writer.writeheader()
        writer.writerows(rows)
